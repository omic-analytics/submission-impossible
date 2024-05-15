import json
import openpyxl

def write_sra_metadata(json_file, bioproject_accession, organism, output_file):
  """
  Reads data from a JSON file and writes it to a user-specified NCBI SRA metadata Excel file,
  preserving column names, sorting entries by barcode number, and using isolation_source from JSON.

  Args:
    json_file: Path to the JSON file containing sample metadata.
    bioproject_accession: User-provided BioProject accession.
    organism: User-provided organism name.
    output_file: User-specified filename for the output Excel file (including .xlsx extension).
  """
  # Open the JSON file and load data
  with open(json_file, 'r') as f:
    data = json.load(f)

  # Prompt user for additional data
  geo_loc_name = "Philippines" # input("Enter Geo Location Name: ")
  host = "Homo sapiens / human" # input("Enter Host: ")
  host_disease = "Covid-19" # input("Enter Host Disease (if any): ")

  # Open the Excel workbook (template)
  wb = openpyxl.load_workbook(filename="SARS-CoV-2.cl.1.0.xlsx")
  sheet = wb.active

  # Define a dictionary to store sample data with barcode as key
  sample_data = {}
  for sample in data:
    barcode = f"{sample.get('barcode', '00')}".zfill(2)  # Handle missing barcode with '00'
    sample_data[barcode] = {
      "sequencing_id": sample["sequencing_id"],
      "dru": sample["dru"],
      "date_of_collection": sample["date_of_collection"],
      "organism": organism,  # Pre-populate organism
      "geo_loc_name": geo_loc_name,  # Pre-populate geo location
      "host": host,  # Pre-populate host
      "host_disease": host_disease,  # Pre-populate host disease
      "sple_type": sample["sple_type"],  # Use JSON value or default
    }

  # Sort sample data by barcode
  sorted_samples = dict(sorted(sample_data.items()))

  # Write sample data starting from row 12 (A12)
  row = 13
  for barcode, sample in sorted_samples.items():
    sheet.cell(row, 1).value = sample["sequencing_id"]
    sheet.cell(row, 3).value = bioproject_accession
    sheet.cell(row, 4).value = sample["organism"]
    sheet.cell(row, 5).value = sample["dru"]
    sheet.cell(row, 6).value = sample["date_of_collection"]
    sheet.cell(row, 7).value = sample["geo_loc_name"]
    sheet.cell(row, 8).value = sample["host"]
    sheet.cell(row, 9).value = sample["host_disease"]
    
    # Construct isolate string
    year = sample["date_of_collection"].split("-")[0]
    isolate = f"{sample['organism']}/{sample['geo_loc_name']}/{sample['sequencing_id']}/{year}"
    sheet.cell(row, 10).value = isolate

    sheet.cell(row, 11).value = sample["sple_type"]

    row += 1

  # Save the workbook to user-specified file
  wb.save(output_file)
  print(f"Data written to {output_file}!")

# Replace with your actual JSON file path
json_file = "test.json"

# Get user input
bioproject_accession = "TEST12345" # input("Enter BioProject Accession: ")
organism = "SARS-CoV-2" # input("Enter Organism Name: ")

# Get user input for output filename
output_file = "test_ncbi_metadata.xlsx" # input("Enter desired filename for the output file (including .xlsx): ")

# Write data to Excel file
write_sra_metadata(json_file, bioproject_accession, organism, output_file)
